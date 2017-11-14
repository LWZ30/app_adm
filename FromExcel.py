# 功能：从Excel读取数据
# 例子：from_excel（“c:\\file.xlsx”）
from openpyxl.reader.excel import load_workbook


def from_excel(file):
    # 读取文件
    wb = load_workbook(file)
    # 获取所有表名
    sheetnames = wb.get_sheet_names()
    # 取第一张表
    ws = wb.get_sheet_by_name(sheetnames[0])
    # 建立存储数据的字典
    data_dic = []
    # 把数据存到字典中
    for rx in range(1, ws.max_row + 1):  # 循环行
        temp_list = []
        for cx in range(1, ws.max_column + 1):  # 循环列
            wc = ws.cell(row=rx, column=cx).value
            temp_list.append(wc)
        data_dic.append(temp_list)
    return data_dic
